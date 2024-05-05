import openpyxl
from openpyxl.styles import Alignment

def formating(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    # Определяем шрифт и цвета заливки
    #шрифт = шрифт (цвет = 'FFFFFFFF')
    #fill = PatternFill (start_color = '000000', end_color = '000000', fill_type = 'solid')

    # Определить параметры выравнивания
    alignment1 = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Разбить текст на слова
    #header = 'Заключить этот текст в заголовок'
    #header_words = заголовок.split()



    for cell in ws[1]:
        #cell.value = header_words[i]
        #cell.font = шрифт
        #cell.fill = заполнить
        cell.alignment = alignment1
        print(cell.value)

    wb.save(path)

